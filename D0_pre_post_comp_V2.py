
# pre.xlsx // post.xlsx // output.xlsx


import openpyxl

wb_pre = openpyxl.load_workbook('pre.xlsx')
wb_post = openpyxl.load_workbook('post.xlsx')
sheet_pre = wb_pre.get_sheet_by_name('All KPIs')
sheet_post = wb_post.get_sheet_by_name('All KPIs')
cell_list = []

# getting unique cell list (pre)

for x in range(2,sheet_pre.max_row+1):
    for y in range(2,sheet_pre.max_row+1):
        if sheet_pre.cell(row=y,column=1).value == sheet_pre.cell(row=x,column=1).value:
            if sheet_pre.cell(row=y,column=1).value not in cell_list:
                cell_list.append(sheet_pre.cell(row=y,column=1).value)
            else:
                continue
        else:
            continue

# getting pre rrc total

i = 0
pre_rrc_sum = []
for cell in cell_list:
    for y in range(2,sheet_pre.max_row+1):
        if sheet_pre.cell(row=y,column=1).value == cell:
            i = i+sheet_pre.cell(row=y,column=6).value
        else:
            continue
    pre_rrc_sum.append(i)
    i = 0
    
# getting pre throughput total

i = 0
pre_thpt_sum = []
n =0
for cell in cell_list:
    for y in range(2,sheet_pre.max_row+1):
        if sheet_pre.cell(row=y,column=1).value == cell:
            if bool(sheet_pre.cell(row=y,column=5).value) == False:
                n = 'N.A'
            else:
                n = n+sheet_pre.cell(row=y,column=5).value
        else:
            continue
    if n == 'N.A':
        pre_thpt_sum.append(n)
    else:
        pre_thpt_sum.append(n/3)
    i = 0
    n = 0

    
postcell_list = []

# getting unique cell list (post)

for x in range(2,sheet_post.max_row+1):
    for y in range(2,sheet_post.max_row+1):
        if sheet_post.cell(row=y,column=1).value == sheet_post.cell(row=x,column=1).value:
            if sheet_post.cell(row=y,column=1).value not in postcell_list:
                postcell_list.append(sheet_post.cell(row=y,column=1).value)
            else:
                continue
        else:
            continue

# getting post rrc total

i = 0
post_rrc_sum = []
for cell in postcell_list:
    for y in range(2,sheet_post.max_row+1):
        if sheet_post.cell(row=y,column=1).value == cell:
            i = i+sheet_post.cell(row=y,column=6).value
        else:
            continue
    post_rrc_sum.append(i)
    i = 0  
    
# getting post throughput total

i = 0
post_thpt_sum = []
n = 0
for cell in postcell_list:
    for y in range(2,sheet_post.max_row+1):
        if sheet_post.cell(row=y,column=1).value == cell:
            if bool(sheet_post.cell(row=y,column=5).value) == False:
                n = 'N.A'
            else:
                n = n+sheet_post.cell(row=y,column=5).value
        else:
            continue
    if n == 'N.A':
        post_thpt_sum.append(n)
    else:
        post_thpt_sum.append(n/3)
    i = 0 
    n = 0

# pasting to excel

wb_op = openpyxl.load_workbook('output.xlsx')
preop = wb_op.get_sheet_by_name('pre')
postop = wb_op.get_sheet_by_name('post')
for x in range(len(cell_list)):
    preop.cell(row=x+2,column=1).value = cell_list[x]
    preop.cell(row=x+2,column=2).value = pre_rrc_sum[x]
    preop.cell(row=x+2,column=3).value = pre_thpt_sum[x]
for x in range(len(postcell_list)):
    postop.cell(row=x+2,column=1).value = postcell_list[x]
    postop.cell(row=x+2,column=2).value = post_rrc_sum[x]
    postop.cell(row=x+2,column=3).value = post_thpt_sum[x]
wb_op.save('output.xlsx')